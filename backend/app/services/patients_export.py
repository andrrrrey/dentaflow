"""Full patient-base export to a single .xlsx workbook.

Produces a three-sheet workbook covering the whole patient base:

* «Пациенты»  — one row per patient: contacts + the 1Denta card fields stored
  in ``Patient.raw_1denta_data`` (medical card no., СНИЛС, ОМС, passport, …).
* «Посещения» — one row per appointment: the full visit / treatment history.
* «Услуги»    — one row per purchased service, expanded from each
  appointment's ``services_data`` array.

Patients and their appointments are read in batches (``_BATCH_SIZE``) so a
large base never loads entirely into the session at once. The CPU-bound
workbook serialisation is pushed to a worker thread by the caller.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.appointment import Appointment
from app.models.patient import Patient

_BATCH_SIZE = 500

# Human-readable Russian labels for appointment statuses (mirrors the UI).
_STATUS_LABELS = {
    "scheduled": "Запланирован",
    "confirmed": "Подтверждён",
    "completed": "Завершён",
    "arrived": "Пришёл",
    "cancelled": "Отменён",
    "canceled": "Отменён",
    "no_show": "Не пришёл",
    "notcome": "Не пришёл",
    "unconfirmed": "Не подтверждён",
}

_SOURCE_LABELS = {
    "telegram": "Telegram",
    "site": "Сайт",
    "call": "Звонок",
    "max": "Max/VK",
    "referral": "Реферал",
}

_PATIENT_TYPE_LABELS = {
    "new": "Новый",
    "regular": "Постоянный",
    "noGroup": "Постоянный",
    "potential": "Потенциальный",
    "refused": "Отказался",
    "refuse": "Отказался",
}

_PATIENT_HEADERS = [
    "ID (1Denta)",
    "ФИО",
    "Телефон",
    "Email",
    "Дата рождения",
    "Пол",
    "№ карты",
    "Источник",
    "Тип пациента",
    "Метки",
    "Кол-во визитов",
    "Последний визит",
    "Выручка, ₽",
    "Средний чек, ₽",
    "Баланс, ₽",
    "Депозит, ₽",
    "LTV",
    "Комментарий",
    "СНИЛС",
    "ИНН",
    "ОМС",
    "Гражданство",
    "Адрес",
    "Паспорт серия",
    "Паспорт номер",
    "Паспорт кем выдан",
    "Дата создания",
]

_VISIT_HEADERS = [
    "ID пациента (1Denta)",
    "ФИО пациента",
    "Телефон",
    "Дата визита",
    "Врач",
    "Услуга",
    "Филиал",
    "Статус",
    "Длительность, мин",
    "Выручка, ₽",
    "Скидка, ₽",
    "Оплачено, ₽",
    "Комментарий",
]

_SERVICE_HEADERS = [
    "ID пациента (1Denta)",
    "ФИО пациента",
    "Дата визита",
    "Услуга",
    "Кол-во",
    "Цена, ₽",
    "Сумма, ₽",
    "Скидка, ₽",
    "Оплачено, ₽",
]


def _fmt_dt(value: datetime | None) -> str:
    return value.strftime("%d.%m.%Y") if value else ""


def _fmt_date(value: date | None) -> str:
    return value.strftime("%d.%m.%Y") if value else ""


def _num(value: Any) -> float | str:
    """Best-effort numeric cast for Excel cells; empty string when unknown."""
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return str(value)


def _gender_label(patient: Patient, card: dict) -> str:
    g = patient.gender or card.get("gender") or card.get("sex")
    if g in ("male", 1, "1"):
        return "М"
    if g in ("female", 2, "2"):
        return "Ж"
    return ""


def _patient_row(patient: Patient) -> list:
    card = patient.raw_1denta_data if isinstance(patient.raw_1denta_data, dict) else {}
    tags = ", ".join(patient.tags) if patient.tags else ""
    return [
        patient.external_id or "",
        patient.name or "",
        patient.phone or "",
        patient.email or "",
        _fmt_date(patient.birth_date),
        _gender_label(patient, card),
        card.get("medical_card") or "",
        _SOURCE_LABELS.get(patient.source_channel or "", patient.source_channel or ""),
        _PATIENT_TYPE_LABELS.get(patient.patient_type or "", patient.patient_type or ""),
        tags,
        _num(card.get("visits_count")),
        _fmt_dt(patient.last_visit_at),
        _num(patient.total_revenue),
        _num(card.get("average_check")),
        _num(card.get("balance")),
        _num(card.get("deposit")),
        _num(patient.ltv_score),
        patient.comment or card.get("comment") or "",
        card.get("snils") or "",
        card.get("inn") or "",
        card.get("oms") or "",
        card.get("citizenship") or "",
        card.get("address") or "",
        card.get("passport_serial") or "",
        card.get("passport_number") or "",
        card.get("passport_issued_by") or "",
        _fmt_dt(patient.created_at),
    ]


def _visit_row(patient: Patient, appt: Appointment) -> list:
    return [
        patient.external_id or "",
        patient.name or "",
        patient.phone or "",
        _fmt_dt(appt.scheduled_at),
        appt.doctor_name or "",
        appt.service or "",
        appt.branch or "",
        _STATUS_LABELS.get(appt.status or "", appt.status or ""),
        appt.duration_min if appt.duration_min is not None else "",
        _num(appt.revenue),
        _num(appt.discount),
        _num(appt.payment_amount),
        appt.comment or "",
    ]


def _service_rows(patient: Patient, appt: Appointment) -> list[list]:
    services = appt.services_data
    if not isinstance(services, list):
        return []
    rows: list[list] = []
    visit_date = _fmt_dt(appt.scheduled_at)
    for s in services:
        if not isinstance(s, dict):
            continue
        rows.append(
            [
                patient.external_id or "",
                patient.name or "",
                visit_date,
                s.get("name") or "",
                _num(s.get("count") if s.get("count") is not None else s.get("amount")),
                _num(s.get("price")),
                _num(s.get("sum") if s.get("sum") is not None else s.get("paySum")),
                _num(s.get("discount")),
                _num(s.get("paySum")),
            ]
        )
    return rows


async def export_all_patients_xlsx(db: AsyncSession) -> bytes:
    """Build the three-sheet patient-base workbook and return it as bytes."""
    import anyio
    from openpyxl import Workbook

    wb = Workbook()
    ws_patients = wb.active
    ws_patients.title = "Пациенты"
    ws_patients.append(_PATIENT_HEADERS)

    ws_visits = wb.create_sheet("Посещения")
    ws_visits.append(_VISIT_HEADERS)

    ws_services = wb.create_sheet("Услуги")
    ws_services.append(_SERVICE_HEADERS)

    offset = 0
    while True:
        patients = list(
            (
                await db.execute(
                    select(Patient)
                    .order_by(Patient.created_at.asc(), Patient.id.asc())
                    .offset(offset)
                    .limit(_BATCH_SIZE)
                )
            )
            .scalars()
            .all()
        )
        if not patients:
            break
        offset += len(patients)

        patient_ids = [p.id for p in patients]
        appts = list(
            (
                await db.execute(
                    select(Appointment)
                    .where(Appointment.patient_id.in_(patient_ids))
                    .order_by(Appointment.scheduled_at.asc())
                )
            )
            .scalars()
            .all()
        )
        by_patient: dict[Any, list[Appointment]] = {}
        for a in appts:
            by_patient.setdefault(a.patient_id, []).append(a)

        for patient in patients:
            ws_patients.append(_patient_row(patient))
            for appt in by_patient.get(patient.id, []):
                ws_visits.append(_visit_row(patient, appt))
                for row in _service_rows(patient, appt):
                    ws_services.append(row)

    buf = BytesIO()
    # Serialisation is CPU-bound — keep it off the event loop.
    await anyio.to_thread.run_sync(wb.save, buf)
    return buf.getvalue()
