"""Слой коррекции ответов для скрипта v2.

Позволяет оператору без программирования «дорабатывать» ответы робота: правка
описывается строкой из 3 столбцов — пример фразы собеседника (trigger), что робот
отвечает сейчас (current_answer, информационно) и правильный ответ (correct_answer).
Когда входящая реплика семантически совпадает с trigger, робот произносит
correct_answer вместо ответа, выданного скриптом. Логика фаз/переходов не меняется.

Источник правды — JSON-файл. Семантический индекс — отдельная ChromaDB-коллекция
(переиспользует тот же стек, что и база знаний).
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from pathlib import Path
from loguru import logger

from app.core.config import get_settings

try:
    import chromadb
    CHROMADB_AVAILABLE = True
except ImportError:
    CHROMADB_AVAILABLE = False
    logger.warning("chromadb not installed — script corrections matching disabled")

# Допустимые значения фазы для правки.
PHASES = ("any", "secretary", "lpr_greeting", "lpr_main", "qualification")

# Заголовки, которые считаем строкой-шапкой таблицы и пропускаем.
_HEADER_HINTS = (
    "пример", "фраза", "секретар", "лпр", "робот", "сейчас",
    "правильн", "нужн", "ответ", "trigger", "current", "correct",
)


def _looks_like_header(cells: list[str]) -> bool:
    """Эвристика: первая строка таблицы похожа на шапку с названиями столбцов."""
    joined = " ".join(c.lower() for c in cells if c)
    hits = sum(1 for h in _HEADER_HINTS if h in joined)
    return hits >= 2


def _normalize_phase(value: str | None) -> str:
    v = (value or "").strip().lower()
    return v if v in PHASES else "any"


def _row_from_cells(cells: list[str]) -> dict | None:
    """Превращает строку таблицы (>=3 ячеек) в правку. None — если строка пустая."""
    cells = [(c or "").strip() for c in cells]
    while len(cells) < 4:
        cells.append("")
    trigger, current_answer, correct_answer, phase = cells[0], cells[1], cells[2], cells[3]
    if not trigger or not correct_answer:
        return None
    return {
        "trigger": trigger,
        "current_answer": current_answer,
        "correct_answer": correct_answer,
        "phase": _normalize_phase(phase),
    }


def _parse_correction_table(file_bytes: bytes, filename: str) -> list[dict]:
    """Парсит таблицу правок из .csv или .xlsx. Возвращает список dict-правок.

    Ожидаемые столбцы: 1) пример фразы, 2) текущий ответ робота, 3) правильный
    ответ, 4) (опц.) фаза. Строка-шапка пропускается автоматически.
    """
    ext = Path(filename).suffix.lower()
    raw_rows: list[list[str]] = []

    if ext == ".csv":
        text = file_bytes.decode("utf-8-sig", errors="replace")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:
            dialect = csv.excel  # запятая по умолчанию
        reader = csv.reader(io.StringIO(text), dialect)
        raw_rows = [list(r) for r in reader]

    elif ext == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            raise ValueError("Для чтения .xlsx требуется библиотека openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(["" if c is None else str(c) for c in row])
        wb.close()

    else:
        raise ValueError(f"Неподдерживаемый формат файла: {ext} (нужен .csv или .xlsx)")

    if not raw_rows:
        return []

    # Пропускаем строку-шапку, только если есть ещё строки данных под ней
    # (одиночная строка почти всегда — данные, а не только заголовок).
    if len(raw_rows) > 1 and _looks_like_header(raw_rows[0]):
        raw_rows = raw_rows[1:]

    rows: list[dict] = []
    for cells in raw_rows:
        parsed = _row_from_cells(cells)
        if parsed:
            rows.append(parsed)
    return rows


class ScriptCorrectionsService:
    """CRUD-хранилище правок (JSON) + семантический матчер (ChromaDB)."""

    COLLECTION_NAME = "script_corrections"

    def __init__(self):
        self.settings = get_settings()
        self.threshold = self.settings.script_correction_threshold
        self._path = Path(self.settings.script_corrections_file)
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._items: list[dict] = self._load()

        self._client = None
        self._collection = None
        if CHROMADB_AVAILABLE:
            try:
                Path(self.settings.knowledge_base_dir).mkdir(parents=True, exist_ok=True)
                self._client = chromadb.PersistentClient(path=self.settings.knowledge_base_dir)
                self._collection = self._client.get_or_create_collection(
                    name=self.COLLECTION_NAME,
                    metadata={"hnsw:space": "cosine"},
                )
                logger.info("Script corrections collection ready")
                self._reindex()
            except Exception as e:
                logger.error(f"Failed to initialize script corrections index: {e}")

    @property
    def index_available(self) -> bool:
        return self._collection is not None

    # ── Персистентность ───────────────────────────────────────────────────────

    def _load(self) -> list[dict]:
        if not self._path.exists():
            return []
        try:
            data = json.loads(self._path.read_text("utf-8"))
            return data if isinstance(data, list) else []
        except Exception as e:
            logger.error(f"Failed to load script corrections: {e}")
            return []

    def _save(self) -> None:
        self._path.write_text(
            json.dumps(self._items, ensure_ascii=False, indent=2), "utf-8"
        )

    def _reindex(self) -> None:
        """Полностью перестраивает ChromaDB-индекс из включённых правок."""
        if not self.index_available:
            return
        try:
            existing = self._collection.get()["ids"]
            if existing:
                self._collection.delete(ids=existing)
            enabled = [i for i in self._items if i.get("enabled", True) and i.get("trigger")]
            if enabled:
                self._collection.add(
                    ids=[i["id"] for i in enabled],
                    documents=[i["trigger"] for i in enabled],
                    metadatas=[
                        {"correct_answer": i["correct_answer"], "phase": i.get("phase", "any")}
                        for i in enabled
                    ],
                )
        except Exception as e:
            logger.error(f"Script corrections reindex error: {e}")

    # ── CRUD ──────────────────────────────────────────────────────────────────

    @staticmethod
    def _normalize(row: dict, base: dict | None = None) -> dict:
        base = base or {}
        return {
            "id": base.get("id") or str(uuid.uuid4()),
            "trigger": str(row.get("trigger", base.get("trigger", ""))).strip(),
            "current_answer": str(row.get("current_answer", base.get("current_answer", ""))).strip(),
            "correct_answer": str(row.get("correct_answer", base.get("correct_answer", ""))).strip(),
            "phase": _normalize_phase(row.get("phase", base.get("phase", "any"))),
            "enabled": bool(row.get("enabled", base.get("enabled", True))),
        }

    def list(self) -> list[dict]:
        return list(self._items)

    def add(self, row: dict) -> dict:
        item = self._normalize(row)
        self._items.append(item)
        self._save()
        self._reindex()
        return item

    def update(self, item_id: str, row: dict) -> dict | None:
        for idx, it in enumerate(self._items):
            if it["id"] == item_id:
                updated = self._normalize(row, base=it)
                updated["id"] = item_id
                self._items[idx] = updated
                self._save()
                self._reindex()
                return updated
        return None

    def delete(self, item_id: str) -> bool:
        before = len(self._items)
        self._items = [it for it in self._items if it["id"] != item_id]
        if len(self._items) == before:
            return False
        self._save()
        self._reindex()
        return True

    def import_rows(self, rows: list[dict], mode: str = "append") -> int:
        """Массово добавляет/заменяет правки. mode: 'append' | 'replace'."""
        if mode == "replace":
            self._items = []
        imported = 0
        for row in rows:
            self._items.append(self._normalize(row))
            imported += 1
        self._save()
        self._reindex()
        return imported

    # ── Семантическое сопоставление ───────────────────────────────────────────

    async def match(self, user_text: str, phase: str) -> str | None:
        """Возвращает correct_answer, если реплика близка к trigger включённой правки."""
        if not self.index_available or not self._items or not user_text.strip():
            return None
        try:
            results = self._collection.query(
                query_texts=[user_text],
                n_results=3,
                include=["distances", "metadatas"],
            )
            dists = results.get("distances", [[]])[0]
            metas = results.get("metadatas", [[]])[0]
            for dist, meta in zip(dists, metas):
                if dist > self.threshold:
                    break  # отсортировано по возрастанию дистанции
                rule_phase = meta.get("phase", "any")
                if rule_phase in ("any", phase):
                    return meta.get("correct_answer") or None
            return None
        except Exception as e:
            logger.error(f"Script corrections match error: {e}")
            return None

    def preview(self, user_text: str, phase: str) -> list[dict]:
        """Отладка: топ-совпадения с дистанциями (для тестового эндпоинта/UI)."""
        if not self.index_available or not user_text.strip():
            return []
        try:
            results = self._collection.query(
                query_texts=[user_text],
                n_results=3,
                include=["distances", "metadatas", "documents"],
            )
            dists = results.get("distances", [[]])[0]
            metas = results.get("metadatas", [[]])[0]
            docs = results.get("documents", [[]])[0]
            out = []
            for dist, meta, doc in zip(dists, metas, docs):
                rule_phase = meta.get("phase", "any")
                out.append({
                    "trigger": doc,
                    "correct_answer": meta.get("correct_answer", ""),
                    "phase": rule_phase,
                    "distance": round(float(dist), 4),
                    "would_fire": dist <= self.threshold and rule_phase in ("any", phase),
                })
            return out
        except Exception as e:
            logger.error(f"Script corrections preview error: {e}")
            return []

    def export_rows(self, fmt: str = "xlsx") -> tuple[bytes, str, str]:
        """Выгружает текущие правки. Возвращает (bytes, content_type, filename)."""
        header = ["Пример фразы собеседника", "Что отвечает сейчас", "Правильный ответ", "Фаза"]
        data = [[i["trigger"], i["current_answer"], i["correct_answer"], i.get("phase", "any")]
                for i in self._items]

        if fmt == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf, delimiter=";")
            writer.writerow(header)
            writer.writerows(data)
            return (
                buf.getvalue().encode("utf-8-sig"),
                "text/csv; charset=utf-8",
                "script_corrections.csv",
            )

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "corrections"
        ws.append(header)
        for r in data:
            ws.append(r)
        out = io.BytesIO()
        wb.save(out)
        wb.close()
        return (
            out.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "script_corrections.xlsx",
        )


# Хелпер для роутов
parse_correction_table = _parse_correction_table
